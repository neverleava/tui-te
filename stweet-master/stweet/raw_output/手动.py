import requests
import json
import jsonsearch
cookies = {
    'guest_id_marketing': 'v1%3A166496725412030245',
    'guest_id_ads': 'v1%3A166496725412030245',
    'personalization_id': '"v1_nfq4C9bGeIs1CiyeCfQ+Ng=="',
    'guest_id': 'v1%3A166496725412030245',
    '_ga': 'GA1.2.1171576380.1664967258',
    '_gid': 'GA1.2.1970596962.1666882427',
    'g_state': '{"i_l":3,"i_p":1667488478386}',
    'ct0': '646a567d77acdd9a01d0179cb204db5c',
    'gt': '1585985831709605888',
}

headers = {
    'authority': 'twitter.com',
    'accept': '*/*',
    'accept-language': 'zh-CN,zh;q=0.9,en;q=0.8,en-GB;q=0.7,en-US;q=0.6,und;q=0.5',
    'authorization': 'Bearer AAAAAAAAAAAAAAAAAAAAANRILgAAAAAAnNwIzUejRCOuH5E6I8xnZz4puTs%3D1Zv7ttfk8LF81IUq16cHjhLTvJu4FA33AGWWjCpTnA',
    # Requests sorts cookies= alphabetically
    # 'cookie': 'guest_id_marketing=v1%3A166496725412030245; guest_id_ads=v1%3A166496725412030245; personalization_id="v1_nfq4C9bGeIs1CiyeCfQ+Ng=="; guest_id=v1%3A166496725412030245; _ga=GA1.2.1171576380.1664967258; _gid=GA1.2.1970596962.1666882427; g_state={"i_l":3,"i_p":1667488478386}; ct0=646a567d77acdd9a01d0179cb204db5c; gt=1585985831709605888',
    'referer': 'https://twitter.com/search?q=Bitcoin%20until%3A2022-11-21%20since%3A2021-01-01&src=typed_query',
    'sec-ch-ua': '"Chromium";v="106", "Microsoft Edge";v="106", "Not;A=Brand";v="99"',
    'sec-ch-ua-mobile': '?0',
    'sec-ch-ua-platform': '"Windows"',
    'sec-fetch-dest': 'empty',
    'sec-fetch-mode': 'cors',
    'sec-fetch-site': 'same-origin',
    'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/106.0.0.0 Safari/537.36 Edg/106.0.1370.52',
    'x-csrf-token': '646a567d77acdd9a01d0179cb204db5c',
    'x-guest-token': '1585985831709605888',
    'x-twitter-active-user': 'yes',
    'x-twitter-client-language': 'zh-cn',
}

params = {
    'include_profile_interstitial_type': '1',
    'include_blocking': '1',
    'include_blocked_by': '1',
    'include_followed_by': '1',
    'include_want_retweets': '1',
    'include_mute_edge': '1',
    'include_can_dm': '1',
    'include_can_media_tag': '1',
    'include_ext_has_nft_avatar': '1',
    'skip_status': '1',
    'cards_platform': 'Web-12',
    'include_cards': '1',
    'include_ext_alt_text': 'true',
    'include_ext_limited_action_results': 'false',
    'include_quote_count': 'true',
    'include_reply_count': '1',
    'tweet_mode': 'extended',
    'include_ext_collab_control': 'true',
    'include_entities': 'true',
    'include_user_entities': 'true',
    'include_ext_media_color': 'true',
    'include_ext_media_availability': 'true',
    'include_ext_sensitive_media_warning': 'true',
    'include_ext_trusted_friends_metadata': 'true',
    'send_error_codes': 'true',
    'simple_quoted_tweet': 'true',
    'q': 'Bitcoin until:2022-11-21 since:2021-01-01',
    'count': '20',
    'query_source': 'typed_query',
    'pc': '1',
    'spelling_corrections': '1',
    'include_ext_edit_control': 'true',
    'ext': 'mediaStats,highlightedLabel,hasNftAvatar,voiceInfo,enrichments,superFollowMetadata,unmentionInfo,editControl,collab_control,vibe',
}

response = requests.get('https://twitter.com/i/api/2/search/adaptive.json', params=params, cookies=cookies, headers=headers)
url_token = "https://api.twitter.com/1.1/guest/activate.json"
def get_token():
    token = json.loads(requests.post(url_token, headers=headers).text)['guest_token']
    headers['x-guest-token'] = token
def parse_Twitter_users(country,keyword,content_json):
    jsondata = JsonSearch(object=content_json, mode='j')
    # channelAboutFullMetadataRenderer
    # print(1)
    rows = []
    user_ids = jsondata.search_all_value(key='users')

    screen_names = jsondata.search_all_value(key='screen_name')
    locations = jsondata.search_all_value(key='location')
    descriptions = jsondata.search_all_value(key='description')
    followers_counts = jsondata.search_all_value(key='followers_count')
    friends_counts = jsondata.search_all_value(key='friends_count')

    cursor_data = jsondata.search_all_value("cursor")
    cursor_data = JsonSearch(object=cursor_data, mode='j')

    cursor_value = jsondata.search_all_value(key='value')

    print(cursor_value)
    for idx,name in enumerate(screen_names):
        print('第{}个数据'.format(idx))
        screen_name = screen_names[idx]
        url = "https://www.twitter.com/"+str(screen_names[idx])
        flag = filter_already_browsId(country, keyword, url)

        if flag:
            with open('already_word/{}_already_browsId.txt'.format(country), 'a+', encoding='utf-8') as fp:
                fp.write(url + '\n')
            location =  locations[idx]
            description = descriptions[idx]
            followers_count = followers_counts[idx]
            friends_count = friends_counts[idx]

            row = [screen_name,url,location,description,followers_count,friends_count]
            rows.append(row)
    print(rows)
    print(len(rows))
    return rows,cursor_value
import os
import time

import openpyxl
from openpyxl import Workbook


def save(location,keyword,rows):
    print('正在保存:{}'.format(rows))
    save_rows = []
    for row in rows:
        row.append(keyword)
        save_rows.append(row)
    print("{}录入成功".format(str(rows)))
    # craw_t = time.strftime('%Y_%m_%d', time.localtime(time.time()))
    # csvfilename = '../../../Datas/{}_youtube_keyword_{}_5000.xlsx'.format(craw_t,keyword)
    csvfilename = '../../../Datas/Twitter/Twitter_{}.xlsx'.format(location)
    '''
    只需要进行追加插入即可
    '''
    if os.path.exists(csvfilename):
        workbook = openpyxl.load_workbook(csvfilename)
    else:
        workbook = Workbook()
    save_file = csvfilename
    worksheet = workbook.active
    # 每个workbook创建后，默认会存在一个worksheet，对默认的worksheet进行重命名
    worksheet.title = "Sheet1"

    max_row = worksheet.max_row

    for r in range(len(save_rows)):
        for c in range(len(save_rows[0])):
            try:
                worksheet.cell(r + max_row, c + 1).value = save_rows[r][c]
            except:
                worksheet.cell(r + max_row, c + 1).value = ''

    workbook.save(filename=save_file)
    print('保存成功')
get_token()
# parse_Twitter_users('China','Uyghur',content_json)