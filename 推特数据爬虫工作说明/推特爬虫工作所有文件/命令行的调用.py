# import subprocess
#
# def run_command_in_cmd(command, working_directory):
#     try:
#         subprocess.run( command, shell=True, check=True, cwd=working_directory)
#     except subprocess.CalledProcessError as e:
#         print(f"命令执行失败：{e}")
#     except Exception as e:
#         print(f"发生错误：{e}")
#
# # 要执行的命令
# command_to_run = r"twarc2 search 'Uyghur' Uyghur0709-0715.jsonl"
#
# # 要执行命令的工作目录
# working_directory = r"C:\Users\12602"
#
# # 调用函数打开命令行窗口并执行命令
# run_command_in_cmd(command_to_run, working_directory)
import openpyxl

# 打开Excel文件
workbook = openpyxl.load_workbook('cleaned2.xlsx')

# 选择要处理的工作表
sheet = workbook.active

# 选择要处理的列，假设是第二列（B列）
column_index = 2  # 第二列的索引是2
column_letter = openpyxl.utils.get_column_letter(column_index)

# 遍历列中的每个单元格
for row in sheet.iter_rows(min_row=2, min_col=column_index, max_col=column_index):
    cell = row[0]
    cell.value = cell.value.replace('\n', '')  # 去掉换行符

# 保存修改后的Excel文件
workbook.save('modified_excel_file.xlsx')
